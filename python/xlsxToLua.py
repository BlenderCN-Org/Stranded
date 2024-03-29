#!/usr/bin/env python
#coding=utf-8

# excel to lua table
# support type: int, bool, float, string, list, struct, list<int>, list<bool>, list<float>, list<string>, list<struct>

import sys
import traceback
import openpyxl
import codecs
import time

TInt = "int"
TList = "list"
TBool = "bool"
TFloat = "float"
TString = "string"
TStruct = "struct"

def ValToKey(val):
	return val.isdecimal() and "[" + val + "]" or val

def CheckInt(data, args = None):
	return data

def CheckBool(data, args = None):
	return "0" == data and "false" or "true"

def CheckFloat(data, args = None):
	return data

def CheckString(data, args = None):
	return "\"" + data.replace("\n", "\\n") + "\""

def CheckList(data, args):
	vals = data.split(";")
	vals.pop()
	if len(vals) == 0:
		return ""
	func = args[0]["func"]
	args = args[0]["args"]
	array = [func(v, args) for k, v in enumerate(vals)]
	return "{ " + ", ".join(array) + " }"

def CheckStruct(data, args):
	vals = data.split(",")
	vals.pop()
	if len(vals) == 0:
		return ""
	array = []
	for k, v in enumerate(vals):
		val = args[k]["func"](v, args[k]["args"])
		name = args[k]["name"]
		if len(val) != 0:
			array.append(name + " = " + val)
	return "{ " + ", ".join(array) + " }"

def GetValue(sheet, row, col):
	return unicode(sheet.cell(row = row, column = col).value or "")

def GetLine(sheet, row):
	return [GetValue(sheet, row, col) for col in xrange(1, sheet.max_column + 1)]

def CheckParses(fields):
	result = []
	for k, field in enumerate(fields):
		pos = field.rfind(":")
		type = field[0:pos]
		name = field[pos+1:]
		parse = { "args": None }
		if TInt == type[:len(TInt)]:
			parse["func"] = CheckInt
		elif TBool == type[:len(TBool)]:
			parse["func"] = CheckBool
		elif TFloat == type[:len(TFloat)]:
			parse["func"] = CheckFloat
		elif TString == type[:len(TString)]:
			parse["func"] = CheckString
		elif TList == type[:len(TList)]:
			parse["func"] = CheckList
			parse["args"] = CheckParses([type[len(TList)+1:-1]])
		elif TStruct == type[:len(TStruct)]:
			parse["func"] = CheckStruct
			parse["args"] = CheckParses(type[len(TStruct)+1:-1].split(","))

		if len(name) == 0:
			parse["name"] = ValToKey(unicode(k))
		else:
			parse["name"] = ValToKey(name)

		result.append(parse)
	return result

def CheckChunk(parses, fields):
	chunk = []
	for index, field in enumerate(fields):
		key = parses[index]["name"]
		val = parses[index]["func"](field, parses[index]["args"])
		if len(val) != 0:
			chunk.append(key + " = " + val)
	return "{ " + ", ".join(chunk) + " }"

def Export(sheet):
	#		first line comment
	#		second line type
	parses = CheckParses(GetLine(sheet, 2))
	#		third conent
	chunks, indexs, parse0 = [], [], parses[0]
	for row in xrange(3, sheet.max_row + 1):
		value = GetValue(sheet, row, 1)
		index = parse0["func"](value, parse0["args"])
		index.append(index)

		chunk = CheckChunk(parses, GetLine(sheet, row))
		chunks.append("%s = %s" % (ValToKey(index), chunk))
		print row, index

	result = [	"return ", \
				"{ " + "\n".join(chunks) + " }", \
				",", \
				"{ " + ", ".join(indexs) + " }", ]
	return "\n".join(result)

def main(opath, ipath):
	xlsx = openpyxl.load_workbook(opath)
	sheet = xlsx.get_sheet_by_name(xlsx.sheetnames[0])
	result = Export(sheet)
	with codecs.open(ipath, "w", "utf-8") as f:
		f.write(result)

if __name__ == '__main__':
	try:
		main("D:\\MyWork\\py\\a.xlsx", "D:\\MyWork\\py\\a.lua")
	except Exception as e:
		traceback.print_exc()
		sys.exit(1)